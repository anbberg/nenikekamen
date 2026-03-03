import datetime as dt
from pathlib import Path

from openpyxl import load_workbook

from excel_log import LOG_HEADERS, append_runs_to_log, _get_or_create_log_sheet, _parse_iso_datetime


def test_parse_iso_datetime_parses_z_and_offset():
    z_ts = "2026-01-01T10:00:00Z"
    offset_ts = "2026-01-01T11:00:00+01:00"
    invalid = "not-a-timestamp"

    parsed_z = _parse_iso_datetime(z_ts)
    parsed_offset = _parse_iso_datetime(offset_ts)
    parsed_invalid = _parse_iso_datetime(invalid)

    assert parsed_z is not None
    assert parsed_offset is not None
    # 10:00Z == 11:00+01:00
    assert parsed_z == parsed_offset
    assert parsed_invalid is None


def test_get_or_create_log_sheet_creates_file_and_headers(tmp_path):
    wb_path = tmp_path / "log.xlsx"
    sheet_name = "Logg"

    ws = _get_or_create_log_sheet(wb_path, sheet_name)

    assert wb_path.exists()
    assert ws.title == sheet_name

    # First row should match LOG_HEADERS for a brand-new sheet
    header_values = [cell.value for cell in ws[1][: len(LOG_HEADERS)]]
    assert header_values == LOG_HEADERS


def test_get_or_create_log_sheet_does_not_overwrite_existing_headers(tmp_path):
    wb_path = tmp_path / "log.xlsx"
    sheet_name = "Logg"

    # First create with our default headers
    ws = _get_or_create_log_sheet(wb_path, sheet_name)
    # Manually change a header to simulate user editing in Excel
    ws["A1"].value = "custom_header"
    ws.parent.save(wb_path)

    # Call helper again; it should NOT overwrite existing headers
    ws2 = _get_or_create_log_sheet(wb_path, sheet_name)
    assert ws2["A1"].value == "custom_header"


def test_append_runs_to_log_writes_expected_values(tmp_path):
    wb_path = tmp_path / "log.xlsx"
    sheet_name = "Logg"

    # One synthetic Strava activity
    run = {
        "id": 123,
        "name": "Test Run",
        "start_date_local": "2026-01-02T09:10:21Z",
        "distance_m": 5000.0,
        "moving_time_s": 1500,  # 25 minutes
        "total_elevation_gain": 42.0,
        "average_heartrate": 140.0,
        "max_heartrate": 170.0,
        "sport_type": "Run",
        "strava_url": "https://www.strava.com/activities/123",
    }

    append_runs_to_log(wb_path, sheet_name, [run])

    wb = load_workbook(wb_path, data_only=True)
    ws = wb[sheet_name]

    # There should be one data row after the header
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    start_dt, distance_km, duration_s, speed_km_h = row[:4]

    # start_datetime is stored as naive datetime
    assert isinstance(start_dt, dt.datetime)
    assert start_dt.year == 2026
    assert start_dt.month == 1
    assert start_dt.day == 2

    # Distance in km rounded to 2 decimals
    assert distance_km == 5.0

    # Raw duration in seconds
    assert duration_s == 1500

    # Speed km/h: distance_km / (seconds / 3600) = 12.0
    assert speed_km_h == 12.0

    # Heart rate and IDs in the right columns
    assert row[4] == 42.0  # elevation_m
    assert row[5] == 140.0  # avg_heart_rate
    assert row[6] == 170.0  # max_heart_rate
    assert row[7] == "Run"  # sport_type
    assert row[8] == 123  # strava_id
    assert row[9] == "https://www.strava.com/activities/123"
    assert row[10] == "Test Run"

