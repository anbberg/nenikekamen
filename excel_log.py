from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


LOG_HEADERS = [
    "start_datetime",
    "week_number",
    "distance_km",
    "duration_s",
    "speed_km_h",
    "elevation_m",
    "avg_heart_rate",
    "max_heart_rate",
    "sport_type",
    "strava_id",
    "strava_url",
    "name",
]


def _get_or_create_log_sheet(wb_path: Path, sheet_name: str) -> Worksheet:
    if wb_path.exists():
        wb = load_workbook(wb_path)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Ensure headers on the first row:
    # - If the sheet is empty, create all standard headers.
    # - If the sheet already has headers, never change existing cells.
    #   We only allow adding new headers to the right if we introduce
    #   more columns in LOG_HEADERS in the future.
    if ws.max_row == 0 or all(cell.value is None for cell in ws[1]):
        # Empty sheet: write all headers.
        for idx, header in enumerate(LOG_HEADERS, start=1):
            ws.cell(row=1, column=idx, value=header)
    else:
        # Existing sheet: respect current headers.
        # If there are fewer header cells than LOG_HEADERS, add missing
        # headers to the right, but never overwrite non-empty cells.
        current_headers = [cell.value for cell in ws[1]]
        existing_cols = len(current_headers)
        for offset, header in enumerate(LOG_HEADERS[existing_cols:], start=1):
            col_idx = existing_cols + offset
            cell = ws.cell(row=1, column=col_idx)
            if cell.value is None:
                cell.value = header

    wb.save(wb_path)
    return ws


def _parse_iso_datetime(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        # Strava returns ISO timestamps, e.g. "2023-01-01T10:00:00Z"
        if value.endswith("Z"):
            value = value.replace("Z", "+00:00")
        return datetime.fromisoformat(value)
    except Exception:
        return None


def _first_empty_data_row(ws: Worksheet) -> int:
    """
    Return the first row index (>= 2) where all log columns are empty.
    If the sheet has no data rows or all rows are filled, return max_row + 1.
    This way we fill gaps when the user has cleared rows instead of deleting them.
    """
    num_cols = len(LOG_HEADERS)
    for row_idx in range(2, ws.max_row + 2):
        if row_idx > ws.max_row:
            return row_idx
        empty = True
        for col_idx in range(1, num_cols + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and val != "":
                empty = False
                break
        if empty:
            return row_idx
    return ws.max_row + 1


def _get_existing_strava_ids(path: Path, sheet_name: str) -> Set[Any]:
    """
    Return a set of Strava IDs that are already present in the log sheet.
    """
    if not path.exists():
        return set()

    wb = load_workbook(path, read_only=True)
    if sheet_name not in wb.sheetnames:
        return set()

    ws = wb[sheet_name]
    ids: Set[Any] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        # Some rows may have fewer columns if the sheet was edited manually.
        # Guard against short tuples before accessing strava_id column.
        if len(row) <= 9:
            continue
        strava_id = row[9]  # strava_id column (0-based)
        if strava_id is not None:
            ids.add(strava_id)
    return ids


def append_runs_to_log(
    path: Path,
    sheet_name: str,
    runs: Iterable[Dict[str, Any]],
) -> None:
    """
    Append Strava run data as rows to the log sheet, skipping runs
    whose Strava ID already exists in the log.
    """
    # Ensure workbook and sheet exist
    _get_or_create_log_sheet(path, sheet_name)
    existing_ids = _get_existing_strava_ids(path, sheet_name)

    wb = load_workbook(path)
    ws = wb[sheet_name]
    next_row = _first_empty_data_row(ws)

    for run in runs:
        strava_id = run.get("id")
        if strava_id in existing_ids:
            continue

        start_dt_local = _parse_iso_datetime(
            run.get("start_date_local") or run.get("start_date")
        )
        distance_km = (run.get("distance_m") or 0.0) / 1000.0
        moving_time_s = int(run.get("moving_time_s") or 0)
        total_elev = run.get("total_elevation_gain")
        avg_hr = run.get("average_heartrate")
        max_hr = run.get("max_heartrate")
        run_type = run.get("sport_type") or run.get("type") or "Run"
        strava_url = run.get("strava_url")
        name = run.get("name")

        # Store raw numeric values: duration in seconds, speed in km/h
        duration_s: Optional[int] = moving_time_s or None
        speed_km_h: Optional[float] = None
        if distance_km > 0 and moving_time_s > 0:
            speed_km_h = round(distance_km / (moving_time_s / 3600.0), 2)

        # Use full local start timestamp (date + time) for Excel.
        start_dt_excel: Optional[datetime] = (
            start_dt_local.replace(tzinfo=None) if start_dt_local else None
        )
        # ISO week as YYYYWW (e.g. 202601) for pivot/aggregate by week.
        week_number: Optional[int] = None
        if start_dt_local:
            y, w, _ = start_dt_local.isocalendar()
            week_number = y * 100 + w

        row = [
            start_dt_excel,
            week_number,
            round(distance_km, 2) if distance_km else None,
            duration_s,
            speed_km_h,
            total_elev,
            avg_hr,
            max_hr,
            run_type,
            strava_id,
            strava_url,
            name,
        ]

        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=next_row, column=col_idx, value=value)
        next_row += 1
        existing_ids.add(strava_id)

    wb.save(path)
